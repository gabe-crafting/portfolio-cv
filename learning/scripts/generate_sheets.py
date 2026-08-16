"""Generates the JavaScript learning spreadsheets in learning/.

    python3 learning/scripts/generate_sheets.py

Writes javascript-essentials.xlsx (reference) and javascript-practice-tracker.xlsx
(progress tracking with live formulas). Re-running overwrites both files, so edit
content.py rather than the spreadsheets if you want changes to survive.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from content import COLUMNS, STATUS_WEIGHTS, TOPICS, TRACKER_ROWS

OUT_DIR = Path(__file__).resolve().parent.parent

BODY = "Arial"
CODE = "Consolas"

INK = "1F2933"
ACCENT = "16324F"
HEADER_FILL = PatternFill("solid", fgColor=ACCENT)
BAND_FILL = PatternFill("solid", fgColor="F2F5F8")
INPUT_FILL = PatternFill("solid", fgColor="FFF9C4")
HAIRLINE = Border(bottom=Side(style="thin", color="D6DDE4"))

TITLE_FONT = Font(name=BODY, size=15, bold=True, color=ACCENT)
SUBTITLE_FONT = Font(name=BODY, size=10, italic=True, color="5A6772")
HEAD_FONT = Font(name=BODY, size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name=BODY, size=10, color=INK)
CODE_FONT = Font(name=CODE, size=9, color="0B3A5B")
BOLD = Font(name=BODY, size=10, bold=True, color=INK)
NOTE_FONT = Font(name=BODY, size=9, italic=True, color="5A6772")

TOPIC_WIDTHS = [26, 44, 50, 28, 60]
WRAP_TOP = Alignment(wrap_text=True, vertical="top")


def put(ws, coord, value, font=None, fill=None, align=None, fmt=None, literal=False):
    """Writes a cell.

    literal=True marks prose and code samples, so a snippet such as '=== strict
    equality' is stored as text instead of being parsed as a formula.
    """
    cell = ws[coord]
    cell.value = value
    if literal and isinstance(value, str) and value.startswith("="):
        cell.data_type = "s"
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    if fmt:
        cell.number_format = fmt
    return cell


def style_topic_sheet(ws, title, subtitle, rows):
    ws.sheet_view.showGridLines = False
    put(ws, "A1", title, TITLE_FONT)
    put(ws, "A2", subtitle, SUBTITLE_FONT)
    put(ws, "A3", "Back to the Start here sheet for the full contents.", NOTE_FONT)

    header_row = 5
    for idx, name in enumerate(COLUMNS, start=1):
        cell = put(ws, f"{get_column_letter(idx)}{header_row}", name, HEAD_FONT, HEADER_FILL)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 22

    for offset, row in enumerate(rows):
        r = header_row + 1 + offset
        band = BAND_FILL if offset % 2 else None
        for idx, value in enumerate(row, start=1):
            font = CODE_FONT if idx in (3, 4) else (BOLD if idx == 1 else BODY_FONT)
            cell = put(ws, f"{get_column_letter(idx)}{r}", value, font, band, WRAP_TOP, literal=True)
            cell.border = HAIRLINE

    for idx, width in enumerate(TOPIC_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:E{header_row + len(rows)}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def build_start_sheet(ws):
    ws.sheet_view.showGridLines = False
    put(ws, "A1", "JavaScript Essentials", Font(name=BODY, size=20, bold=True, color=ACCENT))
    put(ws, "A2", "A reference workbook: one sheet per topic, every row a concept you can look up in seconds.", SUBTITLE_FONT)

    put(ws, "A4", "How to use this workbook", BOLD)
    guide = [
        "Every topic sheet uses the same five columns: Concept, What it does, Example, Result, Notes and gotchas.",
        "The header row has a filter. Filter the Notes column for MUTATES on the Arrays sheet to see which methods change the original.",
        "Read the Example column as real code you can paste into a browser console and run.",
        "Track what you have actually practised in javascript-practice-tracker.xlsx, next to this file.",
        "Content lives in learning/scripts/content.py. Edit there and re-run the generator rather than editing cells.",
    ]
    for i, line in enumerate(guide):
        put(ws, f"A{5 + i}", chr(8226) + "  " + line, BODY_FONT)

    start = 5 + len(guide) + 2
    put(ws, f"A{start - 1}", "Contents", BOLD)
    headers = ["Sheet", "What it covers", "Rows"]
    for idx, name in enumerate(headers, start=1):
        put(ws, f"{get_column_letter(idx)}{start}", name, HEAD_FONT, HEADER_FILL)

    for offset, (title, subtitle, rows) in enumerate(TOPICS):
        r = start + 1 + offset
        band = BAND_FILL if offset % 2 else None
        link = put(ws, f"A{r}", title, Font(name=BODY, size=10, bold=True, color="0563C1", underline="single"), band, WRAP_TOP)
        link.hyperlink = f"#'{title}'!A1"
        cell = put(ws, f"B{r}", subtitle, BODY_FONT, band, WRAP_TOP)
        cell.border = HAIRLINE
        put(ws, f"C{r}", len(rows), BODY_FONT, band, Alignment(horizontal="center", vertical="top"))
        link.border = HAIRLINE

    total_row = start + 1 + len(TOPICS)
    put(ws, f"B{total_row}", "Total concepts", BOLD)
    put(ws, f"C{total_row}", f"=SUM(C{start + 1}:C{total_row - 1})", BOLD, None, Alignment(horizontal="center"))

    put(ws, f"A{total_row + 2}", "Generated from learning/scripts/content.py. Examples are standard JavaScript, no framework or library assumed.", NOTE_FONT)

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 92
    ws.column_dimensions["C"].width = 10


def build_reference():
    wb = Workbook()
    build_start_sheet(wb.active)
    wb.active.title = "Start here"
    for title, subtitle, rows in TOPICS:
        style_topic_sheet(wb.create_sheet(title), title, subtitle, rows)
    path = OUT_DIR / "javascript-essentials.xlsx"
    wb.save(path)
    return path


def build_lists_sheet(ws):
    ws.sheet_view.showGridLines = False
    put(ws, "A1", "Lookup lists", TITLE_FONT)
    put(ws, "A2", "Drives the dropdowns and the Progress formula on the Tracker sheet. Change a weight and every total updates.", SUBTITLE_FONT)

    put(ws, "A4", "Status", HEAD_FONT, HEADER_FILL)
    put(ws, "B4", "Progress weight", HEAD_FONT, HEADER_FILL)
    for i, (status, weight) in enumerate(STATUS_WEIGHTS):
        put(ws, f"A{5 + i}", status, BODY_FONT)
        put(ws, f"B{5 + i}", weight, BODY_FONT, INPUT_FILL, None, "0%")

    put(ws, f"A{5 + len(STATUS_WEIGHTS) + 1}", "Assumption: the weights above are a study convention, not a measurement. Adjust them to taste.", NOTE_FONT)

    put(ws, "D4", "Difficulty", HEAD_FONT, HEADER_FILL)
    for i, level in enumerate(["Easy", "Medium", "Hard"]):
        put(ws, f"D{5 + i}", level, BODY_FONT)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["D"].width = 16


def build_tracker_sheet(ws, last_row):
    ws.sheet_view.showGridLines = False
    put(ws, "A1", "Practice tracker", TITLE_FONT)
    put(ws, "A2", "Fill in the yellow columns as you work. Progress and the Dashboard recalculate on their own.", SUBTITLE_FONT)

    headers = ["Topic", "Reference sheet", "Difficulty", "Status", "Confidence 1-5", "Last reviewed", "Progress", "Exercise to complete", "Notes"]
    for idx, name in enumerate(headers, start=1):
        cell = put(ws, f"{get_column_letter(idx)}4", name, HEAD_FONT, HEADER_FILL)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 26

    for offset, (topic, sheet, difficulty, exercise) in enumerate(TRACKER_ROWS):
        r = 5 + offset
        band = BAND_FILL if offset % 2 else None
        put(ws, f"A{r}", topic, BOLD, band, WRAP_TOP)
        put(ws, f"B{r}", sheet, BODY_FONT, band, WRAP_TOP)
        put(ws, f"C{r}", difficulty, BODY_FONT, INPUT_FILL, Alignment(horizontal="center", vertical="top"))
        put(ws, f"D{r}", STATUS_WEIGHTS[0][0], BODY_FONT, INPUT_FILL, WRAP_TOP)
        put(ws, f"E{r}", None, BODY_FONT, INPUT_FILL, Alignment(horizontal="center", vertical="top"))
        put(ws, f"F{r}", None, BODY_FONT, INPUT_FILL, Alignment(horizontal="center", vertical="top"), "yyyy-mm-dd")
        put(
            ws,
            f"G{r}",
            f"=IFERROR(INDEX(Lists!$B$5:$B$8,MATCH(D{r},Lists!$A$5:$A$8,0)),0)",
            BODY_FONT,
            band,
            Alignment(horizontal="center", vertical="top"),
            "0%",
        )
        put(ws, f"H{r}", exercise, BODY_FONT, band, WRAP_TOP)
        put(ws, f"I{r}", None, BODY_FONT, INPUT_FILL, WRAP_TOP)
        for idx in range(1, 10):
            ws[f"{get_column_letter(idx)}{r}"].border = HAIRLINE

    status_dv = DataValidation(type="list", formula1=f"=Lists!$A$5:$A${4 + len(STATUS_WEIGHTS)}", allow_blank=True)
    status_dv.error = "Pick a status from the list."
    ws.add_data_validation(status_dv)
    status_dv.add(f"D5:D{last_row}")

    conf_dv = DataValidation(type="whole", operator="between", formula1=1, formula2=5, allow_blank=True)
    conf_dv.error = "Confidence is a whole number from 1 to 5."
    ws.add_data_validation(conf_dv)
    conf_dv.add(f"E5:E{last_row}")

    diff_dv = DataValidation(type="list", formula1="=Lists!$D$5:$D$7", allow_blank=True)
    ws.add_data_validation(diff_dv)
    diff_dv.add(f"C5:C{last_row}")

    ws.conditional_formatting.add(
        f"G5:G{last_row}",
        ColorScaleRule(start_type="num", start_value=0, start_color="F8C9C4", mid_type="num", mid_value=0.5, mid_color="FFE9A8", end_type="num", end_value=1, end_color="BFE3C6"),
    )
    ws.conditional_formatting.add(
        f"D5:D{last_row}",
        CellIsRule(operator="equal", formula=['"Confident"'], font=Font(name=BODY, size=10, bold=True, color="1B6E3C")),
    )

    widths = [26, 28, 12, 15, 14, 15, 11, 58, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:I{last_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def build_dashboard(ws, last_row):
    ws.sheet_view.showGridLines = False
    put(ws, "A1", "Dashboard", TITLE_FONT)
    put(ws, "A2", "Every figure is a formula over the Tracker sheet. Nothing here is typed in by hand.", SUBTITLE_FONT)

    rows = [
        ("Topics tracked", f"=COUNTA(Tracker!$A$5:$A${last_row})", "0"),
        ("Overall progress", f"=IFERROR(AVERAGE(Tracker!$G$5:$G${last_row}),0)", "0.0%"),
        ("Average confidence", f'=IFERROR(AVERAGEIF(Tracker!$E$5:$E${last_row},">0"),0)', "0.0"),
        ("Topics never reviewed", f"=COUNTBLANK(Tracker!$F$5:$F${last_row})", "0"),
        ("Next topic to pick up", f'=IFERROR(INDEX(Tracker!$A$5:$A${last_row},MATCH(MIN(Tracker!$G$5:$G${last_row}),Tracker!$G$5:$G${last_row},0)),"All done")', "General"),
    ]
    put(ws, "A4", "Summary", HEAD_FONT, HEADER_FILL)
    put(ws, "B4", "Value", HEAD_FONT, HEADER_FILL)
    for i, (label, formula, fmt) in enumerate(rows):
        r = 5 + i
        put(ws, f"A{r}", label, BODY_FONT)
        put(ws, f"B{r}", formula, BOLD, None, Alignment(horizontal="right"), fmt)
        ws[f"A{r}"].border = HAIRLINE
        ws[f"B{r}"].border = HAIRLINE

    put(ws, "A12", "By status", HEAD_FONT, HEADER_FILL)
    put(ws, "B12", "Topics", HEAD_FONT, HEADER_FILL)
    put(ws, "C12", "Share", HEAD_FONT, HEADER_FILL)
    for i, (status, _weight) in enumerate(STATUS_WEIGHTS):
        r = 13 + i
        put(ws, f"A{r}", status, BODY_FONT)
        put(ws, f"B{r}", f'=COUNTIF(Tracker!$D$5:$D${last_row},$A{r})', BODY_FONT, None, Alignment(horizontal="right"), "0")
        put(ws, f"C{r}", f"=IFERROR(B{r}/$B$5,0)", BODY_FONT, None, Alignment(horizontal="right"), "0.0%")

    put(ws, "A19", "By difficulty", HEAD_FONT, HEADER_FILL)
    put(ws, "B19", "Topics", HEAD_FONT, HEADER_FILL)
    put(ws, "C19", "Confident", HEAD_FONT, HEADER_FILL)
    put(ws, "D19", "Progress", HEAD_FONT, HEADER_FILL)
    for i, level in enumerate(["Easy", "Medium", "Hard"]):
        r = 20 + i
        put(ws, f"A{r}", level, BODY_FONT)
        put(ws, f"B{r}", f'=COUNTIF(Tracker!$C$5:$C${last_row},$A{r})', BODY_FONT, None, Alignment(horizontal="right"), "0")
        put(ws, f"C{r}", f'=COUNTIFS(Tracker!$C$5:$C${last_row},$A{r},Tracker!$D$5:$D${last_row},"Confident")', BODY_FONT, None, Alignment(horizontal="right"), "0")
        put(ws, f"D{r}", f'=IFERROR(SUMIF(Tracker!$C$5:$C${last_row},$A{r},Tracker!$G$5:$G${last_row})/B{r},0)', BODY_FONT, None, Alignment(horizontal="right"), "0.0%")

    put(ws, "A24", "Progress weights come from the Lists sheet, so the totals shift if you change them there.", NOTE_FONT)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16


def build_how_to_use(ws):
    ws.sheet_view.showGridLines = False
    put(ws, "A1", "How to use the tracker", TITLE_FONT)
    put(ws, "A2", "Three sheets: Tracker is where you type, Dashboard is calculated, Lists holds the dropdown values.", SUBTITLE_FONT)

    put(ws, "A4", "Legend", BOLD)
    legend = [
        ("Yellow cells", "The only cells you edit: Difficulty, Status, Confidence, Last reviewed, Notes."),
        ("White cells", "Reference text and formulas. Leave them alone."),
        ("Progress column", "Calculated from Status using the weights on the Lists sheet."),
        ("Dashboard", "Entirely formulas. It updates the moment you change a status."),
    ]
    for i, (term, meaning) in enumerate(legend):
        r = 5 + i
        put(ws, f"A{r}", term, BOLD, INPUT_FILL if i == 0 else None)
        put(ws, f"B{r}", meaning, BODY_FONT, None, WRAP_TOP)

    put(ws, "A11", "Example of a filled-in row", BOLD)
    headers = ["Topic", "Reference sheet", "Difficulty", "Status", "Confidence 1-5", "Last reviewed", "Progress", "Notes"]
    for idx, name in enumerate(headers, start=1):
        put(ws, f"{get_column_letter(idx)}12", name, HEAD_FONT, HEADER_FILL)
    example = ["Arrays: map filter reduce", "05 Arrays", "Medium", "Practising", 3, "2026-08-14", "70%", "reduce still needs the docs, map and filter are fine."]
    for idx, value in enumerate(example, start=1):
        put(ws, f"{get_column_letter(idx)}13", value, BODY_FONT, None, WRAP_TOP)
    put(ws, "A14", "Illustration only. Type your own values on the Tracker sheet, where Progress is a formula rather than text.", NOTE_FONT)

    put(ws, "A16", "A suggested rhythm", BOLD)
    rhythm = [
        "Pick the topic named in Next topic to pick up on the Dashboard.",
        "Read that sheet in javascript-essentials.xlsx, then do the exercise from column H without looking back at it.",
        "Set the status, score your confidence from 1 to 5, and date the row.",
        "Anything scored 3 or below goes back to Learning after a week.",
    ]
    for i, line in enumerate(rhythm):
        put(ws, f"A{17 + i}", chr(8226) + "  " + line, BODY_FONT)

    ws.column_dimensions["A"].width = 30
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 20
    ws.column_dimensions["B"].width = 62


def build_tracker():
    last_row = 4 + len(TRACKER_ROWS)
    wb = Workbook()
    tracker = wb.active
    tracker.title = "Tracker"
    build_tracker_sheet(tracker, last_row)
    build_dashboard(wb.create_sheet("Dashboard"), last_row)
    build_lists_sheet(wb.create_sheet("Lists"))
    build_how_to_use(wb.create_sheet("How to use"))
    path = OUT_DIR / "javascript-practice-tracker.xlsx"
    wb.save(path)
    return path


if __name__ == "__main__":
    for created in (build_reference(), build_tracker()):
        print(f"wrote {created}")
